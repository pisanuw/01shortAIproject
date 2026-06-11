#!/usr/bin/env python3
"""
find_name_conflicts.py - Detect instructor name mismatches between catalog and eval data.

Finds two types of conflicts:
  Rule 1 (eval vs catalog, same section): Same (term, course, section) has different
          instructor names in catalog shards vs IAS eval export.
          Canonical name = eval name (HR-sourced).

  Rule 2 (middle name drop within catalog): Catalog has both "Last,First Middle" and
          "Last,First". Confirmed same person if their professor course histories share
          at least one course code. Ambiguous otherwise.

Already-resolved pairs (left-hand side of fixNames.txt entries) are skipped.

Usage:
  python3 scripts/find_name_conflicts.py             # report only (default)
  python3 scripts/find_name_conflicts.py --apply     # write AUTO entries to fixNames.txt
"""

import logging
import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote, unquote

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT  = SCRIPT_DIR.parent

SHARDS_DIR       = REPO_ROOT / "data" / "shards"
PROFESSORS_DIR   = REPO_ROOT / "data" / "professors"
FIX_NAMES_PATH   = REPO_ROOT / "fixNames.txt"

EVAL_DIR     = REPO_ROOT / "data" / "evals" / "uwb" / "natives" / "0001"
PARAMS_FILE  = EVAL_DIR / "000010.xlsx"

# ── Name helpers (mirrors build_evals.py) ─────────────────────────────────────

def normalize_name(name):
    """Strip nickname-style quotes, e.g. P.V. 'Sundar' -> P.V. Sundar."""
    return re.sub(r"\s'(\w[\w.]*)'", r" \1", name or "").strip()


def make_instructor_id(last, first):
    return quote(f"{normalize_name(last)},{normalize_name(first)}", safe="")


def catalog_name_to_id(name):
    """Convert 'Last,First' catalog name string to URL-encoded professor ID."""
    return quote(name.strip(), safe="")


# ── Load fixNames.txt (already-resolved pairs) ────────────────────────────────

def load_existing_fixes():
    """Return set of left-hand side names (lowercased) from fixNames.txt."""
    if not FIX_NAMES_PATH.exists():
        return set()
    resolved = set()
    for line in FIX_NAMES_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        bad, _ = line.split("=", 1)
        resolved.add(bad.strip().lower())
    return resolved


# ── Load catalog shards ───────────────────────────────────────────────────────

def load_catalog_sections():
    """
    Returns:
      section_map: dict (term, course, section) -> set of catalog instructor names
      all_catalog_names: set of all distinct instructor name strings
    """
    section_map = {}
    all_names = set()
    for campus_dir in SHARDS_DIR.iterdir():
        if not campus_dir.is_dir():
            continue
        for shard_file in campus_dir.glob("*.json"):
            try:
                data = json.loads(shard_file.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError) as exc:
                logger.warning("skipping unreadable shard %s: %s", shard_file, exc)
                continue
            for rec in data.get("records", []):
                instr = (rec.get("instructor") or "").strip()
                if not instr or instr == "TBA":
                    continue
                key = (rec.get("term", ""), rec.get("course", ""), rec.get("section", ""))
                section_map.setdefault(key, set()).add(instr)
                all_names.add(instr)
    return section_map, all_names


# ── Load eval params ──────────────────────────────────────────────────────────

TERM_TO_CODE = {"Autumn": "AUT", "Winter": "WIN", "Spring": "SPR", "Summer": "SUM"}


def load_eval_sections():
    """
    Returns dict (term_code, course, section) -> eval instructor name ("Last,First").
    Uses the primary cross-list row when multiple exist for same EvalID.
    """
    if not PARAMS_FILE.exists():
        return {}
    wb = openpyxl.load_workbook(PARAMS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        last  = str(d.get("LastName")  or "").strip()
        first = str(d.get("FirstName") or "").strip()
        if not last or not first or last.lower() == "unspecified":
            continue
        term  = str(d.get("Term")  or "").strip()
        year  = d.get("Year")
        if not term or not year:
            continue
        term_code = f"{TERM_TO_CODE.get(term, term[:3].upper())}{int(year)}"
        abbrev  = str(d.get("CourseAbbrev")  or "").strip()
        number  = str(d.get("CourseNumber") or "").strip()
        section = str(d.get("Section")      or "").strip()
        course  = f"{abbrev} {number}".strip()
        last_n  = normalize_name(last)
        first_n = normalize_name(first)
        eval_name = f"{last_n},{first_n}"
        key = (term_code, course, section)
        # CrossList=1 is primary; only record primary rows to avoid duplicates
        cross = d.get("CrossList") or 1
        if cross == 1:
            result[key] = eval_name
    return result


# ── Load professor course histories ───────────────────────────────────────────

def load_professor_courses(professor_id):
    """Return set of course codes from data/professors/{id}.json, or empty set."""
    path = PROFESSORS_DIR / f"{professor_id}.json"
    if not path.exists():
        # Try URL-decoded filename (how build_evals.py writes eval-only files)
        decoded = unquote(professor_id)
        path = PROFESSORS_DIR / f"{decoded}.json"
        if not path.exists():
            return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return {r.get("course", "") for r in data.get("records", [])}
    except (json.JSONDecodeError, OSError) as exc:
        logger.warning("could not read %s, treating as empty: %s", path, exc)
        return set()


# ── Rule 1: eval vs catalog section conflict ──────────────────────────────────

def last_name_of(name):
    """Return lowercase last name from 'Last,First' format, or None if invalid."""
    if not name or "," not in name:
        return None
    return name.split(",", 1)[0].strip().lower()


def find_rule1_conflicts(catalog_sections, eval_sections, existing_fixes):
    """
    Returns dict (catalog_name, eval_name) -> list of example section keys.

    Only proposes pairs where:
    - catalog_name is in 'Last,First' format (has a comma)
    - catalog_name and eval_name share the same last name (case-insensitive)
    - When multiple catalog names exist for a section, only the one sharing
      the eval last name is proposed (avoids false positives from co-instructors)
    - catalog_name is not already on the left-hand side of fixNames.txt
    """
    conflicts = {}
    eval_last = last_name_of
    for key, eval_name in eval_sections.items():
        catalog_names = catalog_sections.get(key)
        if not catalog_names:
            continue
        e_last = eval_last(eval_name)
        if not e_last:
            continue
        for cat_name in catalog_names:
            if cat_name.lower() == eval_name.lower():
                continue
            if cat_name.lower() in existing_fixes:
                continue
            # Must be proper Last,First format and share last name
            c_last = last_name_of(cat_name)
            if not c_last or c_last != e_last:
                continue
            pair = (cat_name, eval_name)
            conflicts.setdefault(pair, []).append(key)
    return conflicts


# ── Rule 2: middle name drop within catalog ───────────────────────────────────

def find_rule2_conflicts(all_catalog_names, existing_fixes):
    """
    Returns:
      auto:      list of (long_name, short_name) -- confirmed by course history overlap
      ambiguous: list of (long_name, short_name, reason_str)
    """
    # Group by (last, first_word) to find long/short pairs
    # Name format: "Last,First [Middle...]" or "Last,First"
    groups = {}  # (last_lower, first_lower) -> list of full names
    for name in all_catalog_names:
        if "," not in name:
            continue
        last, rest = name.split(",", 1)
        parts = rest.strip().split()
        if not parts:
            continue
        key = (last.strip().lower(), parts[0].lower())
        groups.setdefault(key, set()).add(name)

    auto = []
    ambiguous = []
    seen_pairs = set()

    for (last_lower, first_lower), names in groups.items():
        if len(names) < 2:
            continue
        # Find the shortest first-name form as the candidate canonical
        by_first_len = sorted(names, key=lambda n: len(n.split(",", 1)[1].strip()))
        short_name = by_first_len[0]
        long_names = by_first_len[1:]

        for long_name in long_names:
            pair = (long_name, short_name)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)

            if long_name.lower() in existing_fixes:
                continue

            # Check course history overlap
            short_id = catalog_name_to_id(short_name)
            long_id  = catalog_name_to_id(long_name)
            short_courses = load_professor_courses(short_id)
            long_courses  = load_professor_courses(long_id)

            if not short_courses and not long_courses:
                ambiguous.append((long_name, short_name,
                                  "no course history found for either"))
                continue

            overlap = short_courses & long_courses
            if overlap:
                auto.append(pair)
            else:
                reason = (
                    f"no shared courses "
                    f"(short: {sorted(short_courses)[:3] or 'none'}, "
                    f"long: {sorted(long_courses)[:3] or 'none'})"
                )
                ambiguous.append((long_name, short_name, reason))

    return auto, ambiguous


# ── Report and apply ──────────────────────────────────────────────────────────

def print_report(rule1, rule2_auto, rule2_ambiguous):
    total_auto = len(rule1) + len(rule2_auto)

    print("=" * 70)
    print(f"NAME CONFLICT REPORT  ({datetime.now().strftime('%Y-%m-%d')})")
    print("=" * 70)

    print(f"\n[AUTO] {len(rule1)} Rule-1 (eval vs catalog, same section)")
    if rule1:
        for (cat_name, eval_name), examples in sorted(rule1.items()):
            ex = examples[0]
            print(f"  {cat_name}={eval_name}")
            print(f"    e.g. {ex[1]} {ex[2]} in {ex[0]}")
    else:
        print("  (none)")

    print(f"\n[AUTO] {len(rule2_auto)} Rule-2 (middle name drop, confirmed by course overlap)")
    if rule2_auto:
        for long_name, short_name in sorted(rule2_auto):
            print(f"  {long_name}={short_name}")
    else:
        print("  (none)")

    print(f"\n[AMBIGUOUS] {len(rule2_ambiguous)} Rule-2 cases needing manual review")
    if rule2_ambiguous:
        for long_name, short_name, reason in sorted(rule2_ambiguous):
            print(f"  {long_name}={short_name}  [{reason}]")
    else:
        print("  (none)")

    print()
    print(f"Total AUTO entries: {total_auto}")
    if total_auto:
        print("Run with --apply to write these to fixNames.txt")
    print()


def apply_fixes(rule1, rule2_auto):
    entries = []
    entries.append(f"\n# --- auto-detected conflicts ({datetime.now().strftime('%Y-%m-%d')}) ---")

    if rule1:
        entries.append("# Rule 1: eval name vs catalog name for same section")
        for (cat_name, eval_name), _ in sorted(rule1.items()):
            entries.append(f"{cat_name}={eval_name}")

    if rule2_auto:
        entries.append("# Rule 2: middle name drop (confirmed by shared courses)")
        for long_name, short_name in sorted(rule2_auto):
            entries.append(f"{long_name}={short_name}")

    block = "\n".join(entries) + "\n"
    with FIX_NAMES_PATH.open("a", encoding="utf-8") as f:
        f.write(block)

    total = len(rule1) + len(rule2_auto)
    print(f"Wrote {total} entries to {FIX_NAMES_PATH.name}")
    print("Next step: npm run build:data && npm run build:evals")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--apply", action="store_true",
                        help="Write AUTO entries to fixNames.txt (default: report only)")
    args = parser.parse_args()

    print("Loading fixNames.txt...")
    existing = load_existing_fixes()

    print("Loading catalog shards...")
    catalog_sections, all_catalog_names = load_catalog_sections()

    print("Loading eval params...")
    eval_sections = load_eval_sections()

    print("Finding Rule-1 conflicts (eval vs catalog)...")
    rule1 = find_rule1_conflicts(catalog_sections, eval_sections, existing)

    print("Finding Rule-2 conflicts (middle name drop)...")
    rule2_auto, rule2_ambiguous = find_rule2_conflicts(all_catalog_names, existing)

    print()
    print_report(rule1, rule2_auto, rule2_ambiguous)

    if args.apply:
        if not rule1 and not rule2_auto:
            print("Nothing to apply.")
        else:
            apply_fixes(rule1, rule2_auto)


if __name__ == "__main__":
    main()
