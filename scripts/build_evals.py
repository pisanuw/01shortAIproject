#!/usr/bin/env python3
"""
build_evals.py - Generate per-instructor evaluation dashboards from IAS export data.

Reads:
  data/evals/uwb/natives/0001/000001.xlsx  - rank codes
  data/evals/uwb/natives/0001/000010.xlsx  - evaluation parameters
  data/evals/uwb/natives/0001/000011.csv   - per-item scores
  data/evals/uwb/natives/0001/000012.csv   - global medians and CEI

Outputs per instructor:
  data/evals/{instructor_id}.json          - structured eval data
  public/evals/{instructor_id}.html        - self-contained dashboard

Outputs globally:
  data/evals/eval_index.json               - list of instructor IDs with eval data

Usage:
  python3 scripts/build_evals.py                           # all instructors
  python3 scripts/build_evals.py --instructor Pisan%2CYusuf  # one instructor only
"""

import argparse
import csv
import html as html_mod
import json
import os
from datetime import datetime, timezone
from urllib.parse import quote, unquote

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT   = os.path.dirname(SCRIPT_DIR)

EVAL_DIR    = os.path.join(REPO_ROOT, "data", "evals", "uwb", "natives", "0001")
RANK_FILE   = os.path.join(EVAL_DIR, "000001.xlsx")
PARAMS_FILE = os.path.join(EVAL_DIR, "000010.xlsx")
ITEMS_FILE  = os.path.join(EVAL_DIR, "000011.csv")
MEDIANS_FILE = os.path.join(EVAL_DIR, "000012.csv")

OUT_JSON_DIR = os.path.join(REPO_ROOT, "data", "evals")
OUT_HTML_DIR = os.path.join(REPO_ROOT, "public", "evals")

TERM_ORDER = {"Autumn": 0, "Winter": 1, "Spring": 2, "Summer": 3}

# ── Helpers ────────────────────────────────────────────────────────────────────

def pf(v):
    """Parse float from cell value; return None if blank or unparseable."""
    try:
        s = str(v).strip()
        return round(float(s), 4) if s else None
    except (ValueError, TypeError):
        return None


def pi(v):
    """Parse int from cell value; return None if blank or unparseable."""
    try:
        s = str(v).strip()
        return int(float(s)) if s else None
    except (ValueError, TypeError):
        return None


def normalize_name(name):
    """Strip nickname-style quotes (e.g. P.V. 'Sundar' -> P.V. Sundar) so IDs match
    those generated from the time schedule catalog, which omits the quotes."""
    import re
    return re.sub(r"\s'(\w[\w.]*)'", r" \1", name or "").strip()


def make_instructor_id(last, first):
    """URL-encode 'Last,First' to match the professor ID format used in the main app."""
    return quote(f"{normalize_name(last)},{normalize_name(first)}", safe="")


def term_sort_key(term, year):
    """Return sort key: newest first (higher academic year, then Autumn before Spring)."""
    q = TERM_ORDER.get(term, 99)
    ay = year if term == "Autumn" else year - 1
    return (-ay, q)

# ── Data loading ───────────────────────────────────────────────────────────────

def load_rank_codes():
    wb = openpyxl.load_workbook(RANK_FILE)
    ws = wb.active
    return {str(r[0]): r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}


def load_eval_params():
    """
    Returns dict: eval_id -> {primary: row_dict, courses: [label, ...]}
    Cross-listed rows share an EvalID; CrossList=1 is the primary.
    """
    wb = openpyxl.load_workbook(PARAMS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    evals = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        eid = d["EvalID"]
        if eid not in evals:
            evals[eid] = {"primary": d, "courses": []}
        abbrev  = (d["CourseAbbrev"]  or "").strip()
        number  = str(d["CourseNumber"] or "").strip()
        section = (d["Section"]       or "").strip()
        label   = f"{abbrev} {number}-{section}".strip()
        cross   = d["CrossList"] or 1
        if cross == 1:
            evals[eid]["primary"] = d
            evals[eid]["courses"].insert(0, label)
        else:
            evals[eid]["courses"].append(label)
    return evals


def load_items():
    """Returns dict: eval_id -> list of item dicts, sorted by ItemPosition."""
    by_eval = {}
    with open(ITEMS_FILE, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            eid = int(row["EvalID"])
            if eid not in by_eval:
                by_eval[eid] = []
            by_eval[eid].append({
                "position":   pi(row["ItemPosition"]),
                "itemId":     row["ItemID"],
                "text":       row["ItemText"],
                "median":     pf(row["Median"]),
                "adjMedian":  pf(row["AdjMedian"]),
                "mean":       pf(row["Mean"]),
                "sd":         pf(row["SD"]),
                "collDecile": pi(row["CollDecile"]),
                "instDecile": pi(row["InstDecile"]),
                "n":          pi(row["N"]),
            })
    for eid in by_eval:
        by_eval[eid].sort(key=lambda x: (x["position"] or 9999))
    return by_eval


def load_medians():
    """Returns dict: eval_id -> global stats dict."""
    result = {}
    with open(MEDIANS_FILE, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            eid = int(row["EvalID"])
            result[eid] = {
                "globalMedian":    pf(row.get("MedianGlobal",    "")),
                "adjGlobalMedian": pf(row.get("AdjMedianGlobal", "")),
                "cei":             pf(row.get("CEI",             "")),
                "nGlobal":         pi(row.get("NGlobal",         "")),
            }
    return result

# ── Professor data for eval-only instructors ──────────────────────────────────

TERM_TO_CODE = {"Autumn": "AUT", "Winter": "WIN", "Spring": "SPR", "Summer": "SUM"}


def load_instructor_courses():
    """Return dict: instructor_id -> list of raw eval-param row dicts (all cross-list rows)."""
    wb = openpyxl.load_workbook(PARAMS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    by_inst = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        last = d.get("LastName") or ""
        first = d.get("FirstName") or ""
        if not last or not first or str(last).lower() == "unspecified":
            continue
        iid = make_instructor_id(last, first)
        by_inst.setdefault(iid, []).append(d)
    return by_inst


def build_professor_records_from_evals(instructor_id, instructor_courses):
    """Build professor-page-style records for an instructor using eval param data."""
    rows = instructor_courses.get(instructor_id, [])
    seen, records = set(), []
    for d in rows:
        term = d.get("Term") or ""
        year = d.get("Year")
        if not term or not year:
            continue
        term_code = f"{TERM_TO_CODE.get(term, term[:3].upper())}{year}"
        abbrev  = (d.get("CourseAbbrev")  or "").strip()
        number  = str(d.get("CourseNumber") or "").strip()
        section = (d.get("Section")       or "").strip()
        course  = f"{abbrev} {number}".strip()
        key = (term_code, course, section)
        if key in seen:
            continue
        seen.add(key)
        dept = abbrev.lower().replace(" ", "")
        records.append({
            "campus":      "B",
            "campusName":  "Bothell",
            "dept":        dept,
            "deptName":    (d.get("DepartmentText") or "").strip(),
            "term":        term_code,
            "course":      course,
            "courseTitle": (d.get("CourseTitle") or "").strip(),
            "section":     section,
        })
    records.sort(key=lambda r: (r["term"], r["course"], r["section"]), reverse=True)
    return records


def write_professor_file(instructor_id, name, records, generated_at):
    """Write data/professors/{id}.json for an eval-only instructor.
    Skips if the file already exists (never overwrite catalog-generated data)."""
    path = os.path.join(REPO_ROOT, "data", "professors", f"{instructor_id}.json")
    if os.path.exists(path):
        return False
    os.makedirs(os.path.dirname(path), exist_ok=True)
    payload = {
        "generatedAt":  generated_at,
        "id":           instructor_id,
        "professor":    name,
        "recordCount":  len(records),
        "records":      records,
        "source":       "eval-only",
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return True


# ── Build instructor records ───────────────────────────────────────────────────

def build_instructors(eval_params, items_by_eval, medians, rank_codes):
    instructors = {}
    for eid, info in eval_params.items():
        p    = info["primary"]
        last = p["LastName"]
        first = p["FirstName"]
        if not last or not first or str(last).lower() == "unspecified":
            continue
        iid = make_instructor_id(last, first)
        if iid not in instructors:
            rank = rank_codes.get(str(p["Rank"]), str(p["Rank"]) if p["Rank"] else "")
            instructors[iid] = {
                "instructor": f"{last}, {first}",
                "instructorId": iid,
                "rank": rank,
                "evaluations": [],
            }
        enroll = p["Enrollment"]      or 0
        q      = p["Questionnaires"]  or 0
        gd     = medians.get(eid, {})
        instructors[iid]["evaluations"].append({
            "evalId":         eid,
            "term":           p["Term"],
            "year":           p["Year"],
            "courses":        info["courses"],
            "courseTitle":    (p["CourseTitle"] or "").strip(),
            "dept":           (p["DepartmentText"] or "").strip(),
            "enrollment":     enroll,
            "questionnaires": q,
            "responseRate":   round(q / enroll, 4) if enroll else None,
            "globalMedian":    gd.get("globalMedian"),
            "adjGlobalMedian": gd.get("adjGlobalMedian"),
            "cei":             gd.get("cei"),
            "nGlobal":         gd.get("nGlobal"),
            "items":           items_by_eval.get(eid, []),
        })
    for iid in instructors:
        instructors[iid]["evaluations"].sort(
            key=lambda e: term_sort_key(e["term"], e["year"])
        )
        # Use the rank from the most recent evaluation.
        newest = instructors[iid]["evaluations"][0]
        newest_eval_id = newest["evalId"]
        if newest_eval_id in eval_params:
            newest_rank_code = str(eval_params[newest_eval_id]["primary"]["Rank"] or "")
            instructors[iid]["rank"] = rank_codes.get(newest_rank_code, newest_rank_code)

    return instructors

# ── HTML template ──────────────────────────────────────────────────────────────
# Uses __PLACEHOLDER__ tokens to avoid conflicts with Python format() and JS template literals.

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>__TITLE__ \u2014 Course Evaluations</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Fraunces:ital,wght@0,700;1,700&family=Space+Grotesk:wght@400;500;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --bg: #f2f4ef;
      --card: #ffffff;
      --ink: #172121;
      --muted: #516262;
      --accent: #cf5c36;
      --accent-soft: #f2d0c3;
      --line: #d8dfd1;
      --green: #2d7a47;
      --green-bg: #e8f5ee;
      --amber: #8a5c00;
      --amber-bg: #fef3e2;
      --red: #b91c1c;
      --red-bg: #fdecea;
      --bar-track: #e4e8e0;
    }
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: "Space Grotesk", sans-serif;
      color: var(--ink);
      background: radial-gradient(circle at 20% 10%, #edf3dc 0%, transparent 35%),
        radial-gradient(circle at 90% 90%, #ffe4d8 0%, transparent 45%), var(--bg);
      min-height: 100vh;
      padding: 32px 16px;
    }
    .app {
      max-width: 1060px;
      margin: 0 auto;
      background: color-mix(in srgb, var(--card) 88%, #f9fbf2 12%);
      border: 1px solid var(--line);
      border-radius: 20px;
      box-shadow: 0 20px 50px rgba(23,33,33,0.12);
      padding: 32px;
    }
    .back-link {
      display: inline-block;
      color: var(--accent);
      text-decoration: none;
      font-weight: 700;
      font-size: 0.88rem;
      margin-bottom: 18px;
    }
    .back-link:hover { text-decoration: underline; }
    h1 {
      font-family: "Fraunces", serif;
      font-size: clamp(1.6rem, 3.5vw, 2.4rem);
      letter-spacing: -0.02em;
      margin-bottom: 6px;
    }
    .rank-badge {
      display: inline-block;
      background: var(--accent-soft);
      color: #6d2f1b;
      padding: 4px 12px;
      border-radius: 999px;
      font-size: 0.82rem;
      font-weight: 700;
      margin-bottom: 28px;
    }
    h2 { font-family: "Fraunces", serif; font-size: 1.2rem; margin: 28px 0 12px; }
    hr.divider { border: none; border-top: 1px solid var(--line); margin: 32px 0; }
    /* Summary table */
    .table-wrap { overflow-x: auto; border: 1px solid var(--line); border-radius: 14px; }
    table { width: 100%; border-collapse: collapse; font-size: 0.87rem; }
    thead th {
      background: #f5f7f3;
      padding: 9px 14px;
      text-align: left;
      font-weight: 700;
      color: var(--muted);
      font-size: 0.76rem;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      border-bottom: 1px solid var(--line);
      white-space: nowrap;
    }
    thead th small { display: block; font-size: 0.7rem; text-transform: none; letter-spacing: 0; font-weight: 400; }
    tbody td { padding: 9px 14px; border-bottom: 1px solid var(--line); vertical-align: middle; }
    tbody tr:last-child td { border-bottom: none; }
    tbody tr:hover { background: #f7faf4; }
    .num { text-align: right; font-variant-numeric: tabular-nums; }
    /* Score chips */
    .chip { display: inline-block; padding: 2px 8px; border-radius: 6px; font-weight: 700; font-size: 0.84rem; }
    .chip-green { background: var(--green-bg); color: var(--green); }
    .chip-amber { background: var(--amber-bg); color: var(--amber); }
    .chip-red   { background: var(--red-bg);   color: var(--red); }
    /* Eval cards */
    .eval-card { border: 1px solid var(--line); border-radius: 14px; margin-bottom: 24px; overflow: hidden; }
    .eval-card-header { padding: 14px 20px; background: #f5f7f3; border-bottom: 1px solid var(--line); }
    .eval-card-header h3 { font-family: "Fraunces", serif; font-size: 1rem; margin-bottom: 4px; }
    .eval-response { font-size: 0.82rem; color: var(--muted); }
    /* OSR + CEI info row */
    .osr-cei-row {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 0;
      border-bottom: 1px solid var(--line);
    }
    .osr-box, .cei-box {
      padding: 16px 20px;
      display: flex;
      gap: 16px;
      align-items: flex-start;
    }
    .osr-box { border-right: 1px solid var(--line); }
    .score-desc { font-size: 0.82rem; color: var(--ink); line-height: 1.5; flex: 1; }
    .score-desc strong { font-weight: 700; }
    .score-block { text-align: center; flex-shrink: 0; min-width: 80px; }
    .score-label { font-size: 0.72rem; font-weight: 700; color: var(--muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px; }
    .score-big { font-family: "Fraunces", serif; font-size: 1.8rem; font-weight: 700; line-height: 1; margin-bottom: 4px; }
    .score-scale { font-size: 0.72rem; color: var(--muted); }
    /* Section controls */
    .section-controls {
      padding: 10px 20px;
      display: flex;
      gap: 8px;
      border-bottom: 1px solid var(--line);
      background: #fafbf8;
    }
    .ctrl-btn {
      background: none;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 4px 12px;
      font-size: 0.78rem;
      font-weight: 700;
      color: var(--muted);
      cursor: pointer;
    }
    .ctrl-btn:hover { background: var(--accent-soft); color: #6d2f1b; border-color: #e8bfad; }
    /* Collapsible sections using <details> */
    .item-section { border-bottom: 1px solid var(--line); }
    .item-section:last-child { border-bottom: none; }
    .item-section > summary {
      padding: 10px 20px;
      font-size: 0.78rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      text-transform: uppercase;
      color: var(--muted);
      cursor: pointer;
      list-style: none;
      display: flex;
      align-items: center;
      gap: 8px;
      user-select: none;
    }
    .item-section > summary::-webkit-details-marker { display: none; }
    .item-section > summary::before { content: "\u25ba"; font-size: 0.65rem; transition: transform 0.15s; }
    .item-section[open] > summary::before { transform: rotate(90deg); }
    .item-section > summary:hover { background: #f5f7f3; }
    .scale-note { font-weight: 400; color: #8aabab; text-transform: none; letter-spacing: 0; font-size: 0.76rem; }
    /* Items table inside section */
    .items-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
    .items-table td { padding: 8px 14px; border-bottom: 1px solid var(--line); vertical-align: middle; }
    .items-table tr:last-child td { border-bottom: none; }
    .items-table tr:hover { background: #f7faf4; }
    .item-text-cell { color: var(--ink); }
    /* Bar */
    .bar-cell { min-width: 160px; }
    .bar-wrap { display: flex; align-items: center; gap: 8px; }
    .bar-track { flex: 1; height: 8px; background: var(--bar-track); border-radius: 4px; overflow: hidden; }
    .bar-fill { height: 100%; border-radius: 4px; }
    .bar-val { font-weight: 700; font-size: 0.84rem; min-width: 32px; text-align: right; }
    .n-cell { text-align: right; color: var(--muted); white-space: nowrap; }
    @media (max-width: 640px) {
      .osr-cei-row { grid-template-columns: 1fr; }
      .osr-box { border-right: none; border-bottom: 1px solid var(--line); }
    }
  </style>
</head>
<body>
<div class="app">
  <a class="back-link" href="../index.html#/professor/__INSTRUCTOR_ID__">\u2190 Back to Course History</a>
  <h1>__INSTRUCTOR_NAME__</h1>
  <p class="rank-badge">__RANK__</p>

  <h2>Summary</h2>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>Term</th>
          <th>Course(s)</th>
          <th>Title</th>
          <th class="num">Enrolled</th>
          <th class="num">Responses</th>
          <th class="num">OSR Median<small>0=lowest; 5=highest</small></th>
          <th class="num">CEI<small>1=lowest; 7=highest</small></th>
        </tr>
      </thead>
      <tbody id="summaryBody"></tbody>
    </table>
  </div>

  <hr class="divider">
  <h2>Evaluations</h2>
  <div id="evalCards"></div>
</div>

<script>
const DATA = __DATA_JSON__;

// Item classification by IAS item ID
var SUMMATIVE_IDS   = {S1:1, S6:1, S13:1, S14:1};
var ENGAGEMENT_IDS  = {S180:1, S181:1, S182:1, S183:1, S184:1};
var EXCLUDE_IDS     = {S186:1, S187:1, S188:1, S189:1, combo:1};
// Everything else (S27-S161) is Standard Formative

function sectionOf(itemId) {
  if (SUMMATIVE_IDS[itemId])  return "summative";
  if (ENGAGEMENT_IDS[itemId]) return "engagement";
  if (EXCLUDE_IDS[itemId])    return "exclude";
  return "formative";
}

// Color thresholds: 0-5 scale
function chipClass05(v) {
  if (v === null || v === undefined) return "";
  return v >= 4.5 ? "chip-green" : v >= 3.5 ? "chip-amber" : "chip-red";
}
function barColor05(v) {
  if (v === null || v === undefined) return "#ccc";
  return v >= 4.5 ? "#2d7a47" : v >= 3.5 ? "#b87333" : "#b91c1c";
}
function barWidth05(v) {
  if (v === null || v === undefined) return 0;
  return Math.min(100, Math.max(0, v / 5 * 100)).toFixed(1);
}

// Color thresholds: 1-7 scale (CEI / engagement items)
function chipClass17(v) {
  if (v === null || v === undefined) return "";
  return v >= 5.5 ? "chip-green" : v >= 4.0 ? "chip-amber" : "chip-red";
}
function barColor17(v) {
  if (v === null || v === undefined) return "#ccc";
  return v >= 5.5 ? "#2d7a47" : v >= 4.0 ? "#b87333" : "#b91c1c";
}
function barWidth17(v) {
  if (v === null || v === undefined) return 0;
  return Math.min(100, Math.max(0, (v - 1) / 6 * 100)).toFixed(1);
}

function fmt(v) { return (v !== null && v !== undefined) ? Number(v).toFixed(2) : "\u2014"; }
function fmtPct(v) { return (v !== null && v !== undefined) ? Math.round(v * 100) + "%" : "\u2014"; }

function chip05(v) {
  if (v === null || v === undefined) return "\u2014";
  return '<span class="chip ' + chipClass05(v) + '">' + fmt(v) + "</span>";
}
function chip17(v) {
  if (v === null || v === undefined) return "\u2014";
  return '<span class="chip ' + chipClass17(v) + '">' + fmt(v) + "</span>";
}

function renderSummary() {
  document.getElementById("summaryBody").innerHTML = DATA.evaluations.map(function(e) {
    return "<tr>" +
      "<td>" + e.term + " " + e.year + "</td>" +
      "<td>" + e.courses.join(" / ") + "</td>" +
      "<td>" + e.courseTitle + "</td>" +
      '<td class="num">' + e.enrollment + "</td>" +
      '<td class="num">' + e.questionnaires + " (" + fmtPct(e.responseRate) + ")</td>" +
      '<td class="num">' + chip05(e.globalMedian) + "</td>" +
      '<td class="num">' + chip17(e.cei) + "</td>" +
      "</tr>";
  }).join("");
}

function renderItemRow(item, isEngagement) {
  var median = item.median;
  var w = isEngagement ? barWidth17(median) : barWidth05(median);
  var color = isEngagement ? barColor17(median) : barColor05(median);
  var cls = isEngagement ? chipClass17(median) : chipClass05(median);
  return "<tr>" +
    '<td class="item-text-cell">' + item.text + "</td>" +
    '<td class="bar-cell">' +
      '<div class="bar-wrap">' +
        '<div class="bar-track"><div class="bar-fill" style="width:' + w + '%;background:' + color + '"></div></div>' +
        '<span class="bar-val ' + cls + '">' + fmt(median) + "</span>" +
      "</div>" +
    "</td>" +
    '<td class="n-cell">' + (item.n !== null ? item.n : "\u2014") + "</td>" +
    "</tr>";
}

function renderSection(items, title, scaleNote, isEngagement, sectionKey) {
  if (!items.length) return "";
  var rows = items.map(function(it) { return renderItemRow(it, isEngagement); }).join("");
  return '<details class="item-section" data-section="' + sectionKey + '">' +
    '<summary>' + title + ' <span class="scale-note">(' + scaleNote + ')</span></summary>' +
    '<table class="items-table"><tbody>' + rows + '</tbody></table>' +
    '</details>';
}

function renderCards() {
  document.getElementById("evalCards").innerHTML = DATA.evaluations.map(function(e) {
    var pct = fmtPct(e.responseRate);
    var osr = e.globalMedian;
    var cei = e.cei;
    var osrClass = chipClass05(osr);
    var ceiClass = chipClass17(cei);

    var summative  = [], engagement = [], formative = [];
    e.items.forEach(function(it) {
      var s = sectionOf(it.itemId);
      if (s === "summative")  summative.push(it);
      else if (s === "engagement") engagement.push(it);
      else if (s === "formative")  formative.push(it);
    });

    var osrHtml = osr !== null
      ? '<div class="score-big ' + osrClass + '">' + fmt(osr) + "</div>"
      : '<div class="score-big">\u2014</div>';
    var ceiHtml = cei !== null
      ? '<div class="score-big ' + ceiClass + '">' + fmt(cei) + "</div>"
      : '<div class="score-big">\u2014</div>';

    return '<div class="eval-card" data-eval-id="' + e.evalId + '">' +
      '<div class="eval-card-header">' +
        '<h3>' + e.term + ' ' + e.year + ' &nbsp;&middot;&nbsp; ' + e.courses.join(' / ') + ' &nbsp;&middot;&nbsp; ' + e.courseTitle + '</h3>' +
        '<p class="eval-response">' + e.questionnaires + ' of ' + e.enrollment + ' students responded (' + pct + ')</p>' +
      '</div>' +
      '<div class="osr-cei-row">' +
        '<div class="osr-box">' +
          '<div class="score-desc"><strong>Overall Summative Rating</strong> represents the combined responses of students to the four global summative items and is presented to provide an overall index of the class\u2019s quality:</div>' +
          '<div class="score-block"><div class="score-label">OSR Median</div>' + osrHtml + '<div class="score-scale">0=lowest; 5=highest</div></div>' +
        '</div>' +
        '<div class="cei-box">' +
          '<div class="score-desc"><strong>Challenge and Engagement Index (CEI)</strong> combines student responses to several IASystem items relating to how academically challenging students found the course to be and how engaged they were:</div>' +
          '<div class="score-block"><div class="score-label">CEI</div>' + ceiHtml + '<div class="score-scale">1=lowest; 7=highest</div></div>' +
        '</div>' +
      '</div>' +
      '<div class="section-controls">' +
        '<button class="ctrl-btn" onclick="toggleAll(this,true)">Expand All</button>' +
        '<button class="ctrl-btn" onclick="toggleAll(this,false)">Collapse All</button>' +
      '</div>' +
      renderSection(summative,  "Summative Items",          "0=lowest; 5=highest", false, "summative") +
      renderSection(engagement, "Student Engagement",       "1=lowest; 7=highest", true,  "engagement") +
      renderSection(formative,  "Standard Formative Items", "0=lowest; 5=highest", false, "formative") +
    '</div>';
  }).join("");
}

function toggleAll(btn, open) {
  var card = btn.closest('.eval-card');
  card.querySelectorAll('details.item-section').forEach(function(d) { d.open = open; });
}

renderSummary();
renderCards();
</script>
</body>
</html>
"""

# ── Output writers ─────────────────────────────────────────────────────────────

def _decoded_filename(instructor_id):
    """Return the URL-decoded name used as the on-disk filename base.
    e.g. 'Pisan%2CYusuf' -> 'Pisan,Yusuf'
    Servers URL-decode request paths, so decoded filenames are found correctly
    when the browser requests the percent-encoded URL."""
    return unquote(instructor_id)


def write_json(data, instructor_id, generated_at):
    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    payload = dict(data, generatedAt=generated_at)
    path = os.path.join(OUT_JSON_DIR, f"{_decoded_filename(instructor_id)}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def write_html(data, instructor_id):
    os.makedirs(OUT_HTML_DIR, exist_ok=True)
    html = HTML_TEMPLATE
    html = html.replace("__TITLE__",           html_mod.escape(data["instructor"]))
    html = html.replace("__INSTRUCTOR_NAME__", html_mod.escape(data["instructor"]))
    html = html.replace("__INSTRUCTOR_ID__",   instructor_id)   # keep encoded for hash link
    html = html.replace("__RANK__",            html_mod.escape(data["rank"]))
    html = html.replace("__DATA_JSON__",       json.dumps(data, ensure_ascii=False))
    path = os.path.join(OUT_HTML_DIR, f"{_decoded_filename(instructor_id)}.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def write_eval_index(instructor_ids, generated_at):
    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    payload = {
        "generatedAt":  generated_at,
        "instructors":  sorted(instructor_ids),
    }
    path = os.path.join(OUT_JSON_DIR, "eval_index.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build eval dashboards from IAS export data.")
    parser.add_argument(
        "--instructor",
        metavar="ID",
        help="URL-encoded instructor ID to build (e.g. Pisan%%2CYusuf). Omit for all.",
    )
    args = parser.parse_args()

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    print("Loading rank codes...")
    rank_codes = load_rank_codes()

    print("Loading evaluation parameters...")
    eval_params = load_eval_params()

    print("Loading per-item scores...")
    items_by_eval = load_items()

    print("Loading global medians...")
    medians = load_medians()

    print("Building instructor data...")
    instructors = build_instructors(eval_params, items_by_eval, medians, rank_codes)

    # ── Add eval-only instructors to the professor index ──────────────────────
    prof_index_path = os.path.join(REPO_ROOT, "data", "professor_index.json")
    try:
        with open(prof_index_path, encoding="utf-8") as f:
            prof_index = json.load(f)
    except FileNotFoundError:
        prof_index = {"professors": []}

    existing_prof_ids = {p["id"] for p in prof_index.get("professors", [])}
    eval_only = {iid: data for iid, data in instructors.items()
                 if iid not in existing_prof_ids}

    if eval_only:
        print(f"Loading instructor course data for {len(eval_only)} eval-only instructor(s)...")
        instructor_courses = load_instructor_courses()
        new_entries = []
        for iid, data in eval_only.items():
            records = build_professor_records_from_evals(iid, instructor_courses)
            # Professor index uses "Last,First" (no space after comma)
            name = data["instructor"].replace(", ", ",", 1)
            created = write_professor_file(iid, name, records, generated_at)
            status = "created" if created else "already exists"
            new_entries.append({
                "id":          iid,
                "name":        name,
                "recordCount": len(records),
                "campuses":    ["Bothell"],
            })
            print(f"  {data['instructor']}: {len(records)} records ({status})")

        # Append only genuinely new entries (guard against re-runs)
        already_in_index = existing_prof_ids
        to_add = [e for e in new_entries if e["id"] not in already_in_index]
        if to_add:
            prof_index["professors"] = prof_index.get("professors", []) + to_add
            with open(prof_index_path, "w", encoding="utf-8") as f:
                json.dump(prof_index, f, ensure_ascii=False, indent=2)
            print(f"  Updated professor_index.json (+{len(to_add)} entries)")

    # ── Filter to one instructor if requested ─────────────────────────────────
    if args.instructor:
        target = args.instructor
        if target not in instructors:
            print(f"Instructor '{target}' not found in eval data.")
            return
        to_build = {target: instructors[target]}
    else:
        to_build = instructors

    # ── Generate JSON and HTML files ──────────────────────────────────────────
    print(f"Generating files for {len(to_build)} instructor(s)...")
    for iid, data in to_build.items():
        json_path = write_json(data, iid, generated_at)
        html_path = write_html(data, iid)
        print(f"  {data['instructor']}: {os.path.basename(json_path)}, {os.path.basename(html_path)}")

    # ── Write eval index based on HTML files that actually exist ──────────────
    # This prevents "View Evaluations" links pointing to non-existent pages.
    os.makedirs(OUT_HTML_DIR, exist_ok=True)
    # Re-encode decoded filenames to get the canonical instructor IDs for the index.
    # e.g. "Pisan,Yusuf.html" -> "Pisan%2CYusuf"
    existing_ids = [
        quote(f[:-5], safe="")
        for f in os.listdir(OUT_HTML_DIR)
        if f.endswith(".html")
    ]
    idx_path = write_eval_index(existing_ids, generated_at)
    print(f"  Eval index ({len(existing_ids)} instructors with pages): {idx_path}")

    print("Done.")


if __name__ == "__main__":
    main()
